import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)

def create_appointments_excel(appointments, start_date=None, end_date=None):
    """
    Создает Excel файл с записями
    
    Args:
        appointments (list): Список словарей с данными записей
        start_date (str): Начальная дата периода (опционально)
        end_date (str): Конечная дата периода (опционально)
        
    Returns:
        str: Путь к созданному файлу
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Записи"
    
    # Заголовок
    headers = [
        "№", "Дата", "Время", "Статус", 
        "Пациент", "Телефон", "Врач", 
        "Источник", "TG ID пользователя"
    ]
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Применяем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        
    # Данные
    for i, apt in enumerate(appointments, 1):
        row_num = i + 1
        
        # Подготовка данных
        status_map = {
            'confirmed': 'Ожидает', 
            'pending': 'Ожидает',
            'visited': 'Посетил',
            'noshow': 'Не пришел',
            'cancelled': 'Отменен'
        }
        
        status_raw = apt.get('status', 'unknown')
        status_text = status_map.get(status_raw, status_raw)
        
        # Цвет статуса
        status_color = "FFFFFF"
        if status_raw in ['confirmed', 'pending']:
            status_color = "EBF1DE" # Зеленоватый
        elif status_raw == 'visited':
            status_color = "DCE6F1" # Голубоватый
        elif status_raw == 'noshow':
            status_color = "F2DCDB" # Красноватый
            
        row_data = [
            i,
            apt.get('appointment_date', ''),
            str(apt.get('appointment_time', ''))[:5],
            status_text,
            apt.get('user_name', ''),
            apt.get('user_phone', ''),
            apt.get('doctor_name', ''),
            "Бот" if apt.get('source') == 'bot' else "Сайт/Другое",
            apt.get('user_telegram_id', '')
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = center_align
            
            # Красим всю строку в зависимости от статуса
            cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

    # Автоширина колонок
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        
    # Сохраняем
    filename = f"appointments_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("data", filename)
    os.makedirs("data", exist_ok=True)
    
    wb.save(filepath)
    return filepath
